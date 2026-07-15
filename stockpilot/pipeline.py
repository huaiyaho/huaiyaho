from __future__ import annotations

import csv
import json
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


ALIASES = {
    "code": {"code", "证券代码", "股票代码", "代码"},
    "name": {"name", "证券名称", "股票名称", "名称"},
    "quantity": {"quantity", "shares", "股票余额", "持仓数量", "证券数量", "可用余额"},
    "cost": {"cost", "cost_price", "成本价", "持仓成本", "成本"},
    "price": {"price", "current_price", "最新价", "当前价", "市价"},
    "market_value": {"market_value", "市值", "证券市值"},
    "profit": {"profit", "浮动盈亏", "持仓盈亏", "盈亏"},
}


@dataclass
class Position:
    code: str
    name: str
    quantity: float
    cost: float
    price: float
    market_value: float
    profit: float
    profit_rate: float


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    text = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text)
    except ValueError:
        return 0.0


def _normalized_key(key: Any) -> str:
    text = str(key).strip()
    lower = text.lower()
    for target, aliases in ALIASES.items():
        if text in aliases or lower in aliases:
            return target
    return lower


def _read_csv(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                return list(csv.DictReader(file))
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Unable to decode CSV: {path}")


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires: pip install openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(item).strip() if item is not None else "" for item in next(rows)]
    return [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]


def read_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def normalize(rows: Iterable[dict[str, Any]]) -> list[Position]:
    positions: list[Position] = []
    for source in rows:
        row = {_normalized_key(key): value for key, value in source.items()}
        code = str(row.get("code", "")).strip().replace(".0", "")
        name = str(row.get("name", "")).strip()
        if not code and not name:
            continue

        quantity = _number(row.get("quantity"))
        cost = _number(row.get("cost"))
        price = _number(row.get("price"))
        market_value = _number(row.get("market_value")) or quantity * price
        profit = _number(row.get("profit")) or quantity * (price - cost)
        principal = quantity * cost
        profit_rate = round((profit / principal * 100) if principal else 0.0, 4)

        positions.append(
            Position(
                code=code.zfill(6) if code.isdigit() and len(code) <= 6 else code,
                name=name,
                quantity=round(quantity, 4),
                cost=round(cost, 4),
                price=round(price, 4),
                market_value=round(market_value, 2),
                profit=round(profit, 2),
                profit_rate=profit_rate,
            )
        )
    return positions


def newest_export(raw_dir: Path) -> Path:
    files = [p for p in raw_dir.iterdir() if p.suffix.lower() in {".csv", ".xlsx", ".xlsm"}]
    if not files:
        raise FileNotFoundError(f"No CSV/XLSX export found in {raw_dir}")
    return max(files, key=lambda item: item.stat().st_mtime)


def write_outputs(positions: list[Position], root: Path, source: Path) -> dict[str, Path]:
    processed = root / "data" / "processed"
    snapshots = root / "data" / "snapshots"
    processed.mkdir(parents=True, exist_ok=True)
    snapshots.mkdir(parents=True, exist_ok=True)

    csv_path = processed / "latest_positions.csv"
    json_path = snapshots / "latest_snapshot.json"
    report_path = snapshots / "latest_report.md"

    fields = list(Position.__dataclass_fields__)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(asdict(position) for position in positions)

    total_value = round(sum(item.market_value for item in positions), 2)
    total_profit = round(sum(item.profit for item in positions), 2)
    snapshot = {
        "version": "0.2.0",
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_file": source.name,
        "summary": {
            "position_count": len(positions),
            "total_market_value": total_value,
            "total_profit": total_profit,
        },
        "positions": [asdict(position) for position in positions],
    }
    json_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# StockPilot Latest Portfolio Report",
        "",
        f"- Source: `{source.name}`",
        f"- Positions: {len(positions)}",
        f"- Total market value: {total_value:,.2f}",
        f"- Total floating P/L: {total_profit:,.2f}",
        "",
        "| Code | Name | Quantity | Cost | Price | Market Value | P/L | Return |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for item in positions:
        lines.append(
            f"| {item.code} | {item.name} | {item.quantity:g} | {item.cost:.4f} | "
            f"{item.price:.4f} | {item.market_value:,.2f} | {item.profit:,.2f} | {item.profit_rate:.2f}% |"
        )
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return {"csv": csv_path, "json": json_path, "report": report_path}


def run(root: Path) -> dict[str, Path]:
    source = newest_export(root / "data" / "raw")
    positions = normalize(read_rows(source))
    if not positions:
        raise ValueError("No valid positions were recognized. Check the exported column names.")
    return write_outputs(positions, root, source)
